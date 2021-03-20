import openpyxl
import re
import docx
import os
'''提取关系'''


def readfromdoc(infile):
    '''从文本中读取句子'''
    doc = docx.Document(infile)
    sentences = []
    for p in doc.paragraphs:
        sentences.extend(filter(lambda x:x,p.text.strip().split('。')))
    return sentences

def Extract_connection_yuanyin(path,outfile):
    '''提取 ‘使、因为、原因’关系 '''

    def regular_extract_shi(sentence):
        text1, text2, text3 = None, None, None
        matchobj = re.search(r'(.+)(使)(.+)', sentence)
        if matchobj != None and matchobj.group(3)[0] != '用':
            text1 = matchobj.group(1)
            text2 = '使'
            text3 = matchobj.group(3)
            # print(text1)
            if len(text1) > 2:
                matchobj1 = re.search(r'(.*)，(.*)', text1[-3:])
                if matchobj1 != None:
                    text1 = text1[:-3] + matchobj1.group(1)
                    text2 = matchobj1.group(2) + text2
                    # print(text2)
                # print(text3)
            if text3[0] in ['之', '致', '得']:
                text2 += text3[0]
                text3 = text3[1:]
                # print(text2)
        return text1, text2, text3

    def regular_extract_yinwei(sentence, lastsentence):
        text1, text2, text3 = None, None, None
        matchobj = re.search(r'(.+)因为(.+)', sentence)
        if matchobj != None:
            text1 = matchobj.group(1)
            text2 = '因为'
            text3 = matchobj.group(2)
            if len(text1) < 3:
                text2 = text1 + text2
                text1 = lastsentence
            if len(text1) > 2:
                matchobj1 = re.search(r'(.*)，(.*)', text1[-3:])
                if matchobj1 != None:
                    text1 = text1[:-3] + matchobj1.group(1)
                    text2 = matchobj1.group(2) + text2
            # print(text3)
        return text1, text2, text3

    def regular_extract_yuanyin(sentence):
        text1, text2, text3 = None, None, None
        matchobj = re.search(r'(.+)原因是(.+)', sentence)
        if matchobj != None:
            text1 = matchobj.group(1)
            text2 = '原因是'
            text3 = matchobj.group(2)
            # print(sentence)
        return text1, text2, text3

    def regular_extract_youyu(sentence, lastsentence):
        text1, text2, text3 = None, None, None
        matchobj = re.search(r'(.+)由于(.+)', sentence)
        if matchobj != None:
            text1 = matchobj.group(1)
            text2 = '由于'
            text3 = matchobj.group(2)
            if matchobj.group(1)[-1] == '是':
                text2 = text1[-1] + text2
                text1 = text1[:-1]
        return text1, text2, text3
    try:
        wb = openpyxl.load_workbook(outfile)
    except:
        wb = openpyxl.Workbook()

    for dir in os.listdir(path):
        if os.path.splitext(dir)[1] == '.docx':
            infile=dir
            sentences = readfromdoc(path+'\\'+infile)

            try:
                ws=wb['使']
            except:
                ws= wb.create_sheet(title='使')
            i=1
            while ws.cell(row=i,column=1).value!=None:
                i+=1
            lastsentence = ''
            for sentence in sentences:
                text1,text2,text3=regular_extract_shi(sentence)
                if text1!=None:
                    ws.cell(row=i, column=1, value=lastsentence)
                    ws.cell(row=i,column=2,value=text1)
                    ws.cell(row=i,column=3,value=text2)
                    ws.cell(row=i,column=4,value=text3)
                    ws.cell(row=i,column=5,value=os.path.splitext(dir)[0])
                    i+=1
                lastsentence=sentence

            try:
                ws = wb['因为']
            except:
                ws = wb.create_sheet(title='因为')
            i = 1
            while ws.cell(row=i, column=1).value != None:
                i += 1
            lastsentence = ''
            for sentence in sentences:
                text1, text2,text3= regular_extract_yinwei(sentence,lastsentence)
                if text1!=None:
                    ws.cell(row=i, column=1, value=lastsentence)
                    ws.cell(row=i, column=2, value=text1)
                    ws.cell(row=i, column=3, value=text2)
                    ws.cell(row=i, column=4, value=text3)
                    ws.cell(row=i, column=5, value=os.path.splitext(dir)[0])
                    i+=1
                lastsentence=sentence

            try:
                ws = wb['原因']
            except:
                ws = wb.create_sheet(title='原因')
            i = 1
            while ws.cell(row=i, column=1).value != None:
                i += 1
            lastsentence = ''
            for sentence in sentences:
                text1, text2, text3 = regular_extract_youyu(sentence,lastsentence)
                if text1 != None:
                    ws.cell(row=i, column=1, value=lastsentence)
                    ws.cell(row=i, column=2, value=text1)
                    ws.cell(row=i, column=3, value=text2)
                    ws.cell(row=i, column=4, value=text3)
                    ws.cell(row=i, column=5, value=os.path.splitext(dir)[0])
                    i += 1
                lastsentence = sentence
            for sentence in sentences:
                text1, text2, text3 = regular_extract_yuanyin(sentence)
                if text1 != None:
                    ws.cell(row=i, column=1, value=lastsentence)
                    ws.cell(row=i, column=2, value=text1)
                    ws.cell(row=i, column=3, value=text2)
                    ws.cell(row=i, column=4, value=text3)
                    ws.cell(row=i, column=5, value=os.path.splitext(dir)[0])
                    i += 1

    wb.save(outfile)


def Extract_connection_baokuo(path,outfile):
    '''提取 ‘包括’关系 '''
    def regular_extract_baokuo(sentence):
        text1, text2, text3 = None, None, None
        matchobj = re.search(r'(.+)①(.+)', sentence)
        if matchobj != None:
            # print(sentence)
            tunc = re.split(r'[①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯]', sentence)
            text1 = tunc[0]
            text2 = '包括'
            text3 = tunc[1:]
        if matchobj == None:
            '''(a)用目前的语料集很少，错误率也挺高的'''
            matchobj = re.search(r'[\(（]a[\)）](.+)', sentence)
            if matchobj != None:
                tunc = re.split(r'[abcd]', sentence)
                text1 = tunc[0][:-1]
                text2 = '包括'
                text3 = []
                for i in range(1, len(tunc)):
                    text3.append(tunc[i][1:-1])
                # print(sentence)
        if matchobj == None:
            '''a)用目前的语料集其实根本提取不出来'''
            matchobj = re.search(r' a[\)）](.+)', sentence)
            if matchobj != None:
                tunc = re.split(r'[abcd]', sentence)
                text1 = tunc[0]
                text2 = '包括'
                text3 = []
                for i in range(1, len(tunc)):
                    text3.append(tunc[i][1:])
        if matchobj == None:
            '''这样做准确律还是太低'''
            # matchobj = re.search(r'[^(（0-9a-zA-Z]1[\)）](.*)[^(（]2[\)）](.+)', sentence)
            '''只能提取几个特例'''
            matchobj = re.search(r'[;；:：]1[\)）](.+)', sentence)
            if matchobj != None:
                tunc = re.split(r'[;；:：]', sentence)
                text1 = tunc[0]
                text2 = '包括'
                text3 = tunc[1:]
        # if matchobj==None:
        '''破折号大部分后面都是对前面的解释，不是包括关系，故不要'''
        #    matchobj = re.search(r'(.+)——(.+)', sentence)
        #    if matchobj!=None:
        #        print(sentence)
        return text1, text2, text3
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(title='包括')
    i=1
    for dir in os.listdir(path):
        if os.path.splitext(dir)[1] == '.docx':
            infile = dir
            sentences = readfromdoc(path + '\\' + infile)
            lastsentence=''
            for sentence in sentences:
                text1, text2, text3 = regular_extract_baokuo(sentence)
                if text1 != None:
                    for text in text3:
                        ws.cell(row=i,column=1,value=lastsentence)
                        ws.cell(row=i, column=2, value=text1)
                        ws.cell(row=i, column=3, value=text2)
                        ws.cell(row=i, column=4, value=text)
                        ws.cell(row=i, column=5, value=os.path.splitext(dir)[0])
                        i += 1
                lastsentence=sentence
    wb.save(outfile)
Extract_connection_yuanyin(r'..\医学相关\医学相关docx',r'../关系提取/原因.xlsx')
#Extract_connection_baokuo(r'..\医学相关\医学相关docx',r'../关系提取/包括.xlsx')