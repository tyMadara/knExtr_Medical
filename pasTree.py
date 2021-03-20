import os
from passage_extract import*
import openpyxl
import pickle

class Node(object):
    def __init__(self,name='',par=None):
        self.child=None
        self.parent=par
        self.name=name
        self.text=''
        self.addr=self
        self.count=0
        self.height=0
        self.index=None
        self.unitylist=[]
    def Addchild(self, name,text=''):
        tmpnode = Node(name,self)
        tmpnode.height=self.height+1
        tmpnode.index=self.count
        tmpnode.text=text
        if self.child!=None:
            self.child.append(tmpnode)
            self.count+=1

        else:
            self.child=[tmpnode]
            self.count=1
        return
    def Getchild(self,i:int):
        if self.count==0:
            return None
        else:
            return self.child[i]
    def show(self):
        print(self.name)
        for i in range(self.count):
            self.child[i].show()
        return
    def passage(self):
        str=''
        str+='\t'*self.height+self.name+'\n'
        textlist=self.text.split('\n')[:-1]
        for i in range(len(textlist)):
            textlist[i]='\t'*self.height+textlist[i]+'\n'
            str+=textlist[i]
        for i in range(self.count):
            str+=self.child[i].passage()
        return str
    def Write2txt(self, outfile):
        with open(outfile,'w',encoding='utf-8') as ofh:
            ofh.write(self.passage())
    def first_leafnode(self):
        node=self
        while node.Getchild(0)!=None:
            node=self.Getchild(0)
        return node
    def last_leafnode(self):
        node = self
        while node.Getchild(-1) != None:
            node = self.Getchild(-1)
        return node
    def Nextnode(self):
        if self.Getchild(0)!=None:
            return self.Getchild(0)
        elif self.parent!=None:
            node=self
            while node.parent!=None and node.index==node.parent.count-1:
                tmp_index=node.index
                node=node.parent
            if node==None:
                return None
            else:
                return node.parent.Getchild(node.index+1)
            #if self.index<self.parent.count-1:
            #    return self.parent.Getchild(self.index+1)

    def make_unitylist(self,infile):
        uList = ulist_Readfromtxt(infile)

        '''提取句子中的实时需要实体列表和实体名最大长度，为了加快运算速度，这里先将这两个部分处理好，避免每次迭代中重复运行。'''
        dict=[]
        for ele in uList:
            dict.append(ele.name)
        max_len = max(len(ele) for ele in dict)

        self._make_unitylist(uList,max_len,dict)

    def _make_unitylist(self, uList,max_len,dict):
        sentences=re.split('。\n',self.text)
        if sentences[-1]=='':
            sentences.pop()
        '''使用贪心算法，当出现‘病理学’时只统计‘病理学’而不统计‘病理’'''
        for sentence in sentences:
            sen_len = len(sentence)
            loc = 0
            length = min(max_len, sen_len - loc)
            while (sen_len != loc):
                if length == 1:
                    # print(line[loc]+'\tO\n')
                    loc += 1
                    length = min(max_len, sen_len - loc)
                elif sentence[loc:loc + length] in dict:
                    self.unitylist.append(unity(sentence[loc:loc + length]))
                    loc += length
                    length = min(max_len, sen_len - loc)
                else:
                    length -= 1

        for i in range(self.count):
            self.Getchild(i)._make_unitylist(uList,max_len,dict)
    def Show_unitylist(self):
        if self.text!='':
            print(self.text)
        PrintuList(self.unitylist)
        for i in range(self.count):
            self.Getchild(i).Show_unitylist()

def make_cataTree(infile):
    '''利用文章目录生成一个目录树'''
    cataTree=Node(os.path.splitext(os.path.basename(infile))[0])
    node=cataTree
    with open(infile,'r',encoding='utf-8') as ifh:
        levellist=[-1]
        for line in ifh.readlines():
            i=0
            while line[i]=='\t':
                i+=1
            t_number = i
            for i in range(len(levellist)-1,-1,-1):
                if levellist[i]>t_number:
                    node=node.parent
                    levellist=levellist[:i]
                elif levellist[i]==t_number:
                    node=node.parent
                    node.Addchild(line[t_number:].strip())
                    node=node.Getchild(-1)
                    levellist=levellist[:i+1]
                    break
                else:
                    levellist.append(t_number)
                    node.Addchild(line[t_number:].strip())
                    node=node.Getchild(-1)
                    break
    return cataTree
def make_pasTree(catafile,infile):
    cataTree = make_cataTree(catafile)
    '''利用目录树和整篇文章生成文章树'''
    with open(infile, 'r', encoding='utf-8') as ifh:
        node=cataTree
        nextnode=node.Nextnode()
        text=''
        for line in ifh.readlines():
            if nextnode==None or line.strip()!=nextnode.name:
                text+=line
            else:
                node.text=text
                node=nextnode
                nextnode=node.Nextnode()
                text=''
        node.text=text
    return cataTree

def Addcata2con(infile):
    '''没写完，先放着'''
    ofh=openpyxl.Workbook()
    ofhs=ofh.create_sheet('包括关系')
    wb=openpyxl.load_workbook(infile)
    ws=wb['包括关系']
    num=1
    while ws.cell(row=num, column=4).value!=None:
        num+=1
    book = ws.cell(row=1, column=4).value
    i=1
    for i in range(num+1):
        if book!=ws.cell(row=i,column=4).value:
            book=ws.cell(row=i,column=4).value
            pasTree=make_pasTree(r'..\医学相关\目录结构'+os.path.basename(os.path.splitext(book)[0])+'.txt',r'..\医学相关\医学相关docx'+os.path.basename(book))
        node=pasTree
        while node!=None and re.search(str(ws.cell(row=i,column=1).value),node.text)!=None:
            node=node.Nextnode()
        j=1
        while node!=None:
            ofhs.cell(row=i,column=j,value=node.name)
            j+=1
            node=node.parent
    ofh.save(r'..\关系提取\节点信息')



#pastree=make_pasTree(r'..\医学相关\目录结构\病理学目录.txt',r'..\医学相关\医学相关txt\病理学.txt')
#pastree.make_unitylist(r'..\实体提取\病理学_unitylist.txt')
#with open(r'..\实体提取\病理学_结构.txt', 'wb') as f:
#    pickle.dump(pastree,f)
#with open(r'..\实体提取\病理学_结构.txt', 'rb') as f:
#    pastree = pickle.load(f)
#pastree.Show_unitylist()




