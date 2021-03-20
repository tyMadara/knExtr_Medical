import re
import jieba.posseg as posseg
import jieba
import docx
import os
import csv
import openpyxl
import codecs
from tqdm import tqdm

class unity(object):
    def __init__(self, uName='',rule=''):
        self.proList = []
        self.addr = self
        self.name = uName
        self.rule=rule

    def addProperty(self, proName):
        self.proList.append(proName)

    def modName(self, uName):
        self.name = uName

def sentence_Readfromtxt(infile):
    '''从文本中读取句子'''
    with open(infile, 'r', encoding='utf-8') as ifh:
        sentences=[]
        for line in ifh.readlines()[:]:
            sentences.extend(filter(lambda x:x,line.strip().split('。')))
    #for i in range(len(sentences)):
        #sentences[i]+='。'
        #print(sentences[i])
    return sentences

def ulist_Readfromtxt(inuList):
    '''用实体列表文档unitylist.txt中的实体构建一个uList'''
    uList=[]
    try:
        with open(inuList,'r',encoding='utf-8') as ifh:
            for line in ifh.readlines():
                matchobj=re.search(r'(.*)\((.*)\)',line)
                if matchobj!=None:
                    tmpuList=unity(matchobj.group(1).strip())
                    tmpuList.proList.extend(matchobj.group(2).split(','))
                    uList.append(tmpuList)
                else:
                    uList.append(unity(line.strip()))
    except:
        uList=[]
    #PrintuList(uList)
    return uList

def sentence_Readfromdoc(infile):
    '''从文本中读取句子'''
    doc = docx.Document(infile)
    sentences = []
    for p in doc.paragraphs:
        sentences.extend(filter(lambda x:x,p.text.strip().split('。')))
    return sentences

def dict_readformuList(inuList):
    '''用实体列表文档unitylist.txt中的实体构建一个dict(这是一个列表)'''
    dict = []
    with open(inuList, 'r', encoding='utf-8') as ifh:
        for line in ifh.readlines():
            matchobj = re.search(r'(.*)\((.*)\)', line)
            if matchobj != None:
                dict.append(matchobj.group(1).strip())
            else:
                dict.append(line.strip())
    # PrintuList(uList)
    return dict


def PrintuList(uList):
    print('[',end='')

    for i in range(len(uList)):
        if len(uList[i].proList) > 0:
            print(uList[i].name, end=' ')
            if i==len(uList):
                print('('+' '.join(uList[i].proList)+')]')
            else:
                print('(' + ' '.join(uList[i].proList) + ')')
        else:
            if i == len(uList)-1:
                print(uList[i].name+']')
            else:
                print(uList[i].name)


def ulist_Addunity(uList, unity):
    '''将一个实体添加到uList里，同时合并属性（若已有）'''
    flag=0
    for ele in uList:
        if ele.name==unity.name:
            for pro in unity.proList:
                if pro not in ele.proList:
                    ele.proList.append(pro)
            flag=1
    if flag==0:
        uList.append(unity)
    return uList

def ulisttxt_Writefromulist(resList, outfile):
    '''把实体列表中的结果添加到实体列表.txt里'''
    uList=ulist_Readfromtxt(outfile)
    for ele in resList:
        uList=ulist_Addunity(uList, ele)
    #print(len(uList))
    with open(outfile,'w',encoding='utf-8') as ofh:
        for ele in uList:
            ofh.write(ele.name+' ')
            if len(ele.proList)>0:
                ofh.write('(')
                ofh.write(','.join(ele.proList))
                ofh.write(')')
            ofh.write('\n')

def dictfile_Writefromulist(uList, file):
    '''用实体列表生成一个字典'''
    dict=[]
    if file!='':
        try:
            with open(file,'r',encoding='utf-8') as ifh:
                for line in ifh.readlines():
                    matchobj=re.search(r'(.*)n',line)
                    if matchobj!=None:
                        #print(matchobj.group(1).split())
                        dict.append(matchobj.group(1).strip())
        except:
            with open(file, 'w', encoding='utf-8') as ofh:
                pass
    for ele in uList:
        if ele.name not in dict and len(ele.name)<10:
            dict.append(ele.name)
    if file=='':
        file=r'dict.txt'
    with open(file,'w',encoding='utf-8') as ofh:
        for ele in dict:
            ofh.write(ele+' n')
            ofh.write('\n')

def gen2list(seg):
    '''将词性标注结果转化为词性列表'''
    resList=[]
    for ele in seg:
        resList.append([ele.word, ele.flag])
    return resList

def poslist_Show(posList):
    '''输出词性列表'''
    for ele in posList:
        print(ele[0]+ele[1],end=' ')
    print('')

def listmerge(list, label_list):
    '''若一个词的标签属于标签列表，则将其标签更改为标签列表的第一个标签，然后将连续的这些标签合并'''
    flag = 0
    resList = []
    for ele in list:
        if ele[1] not in label_list:
            flag = 0
            resList.append([ele[0], ele[1]])
        elif ele[1] in label_list:
            if flag == 0:
                flag = 1
                resList.append([ele[0], label_list[0]])
            else:
                resList[-1][0] = resList[-1][0] + ele[0]
    return resList


def poslist2str(poslist):
    '''将词性列表中的词拼接成字符串'''
    str=''
    for ele in poslist:
        str+=ele[0]
    return str


def greedmatch(text_list,dict,merge=2,jiange=0):
    '''利用贪婪算法和实体列表分割文本'''
    '''输入：句列表，输出：实体列表的列表'''
    '''merge=0:一般模式'''
    '''merge=1:合并相连实体'''
    '''merge=2:合并间隔不大于1的实体'''
    def match(text,max_len,dictlist):
        uList=[]
        text_len = len(text)
        loc = 0
        length = min(max_len, text_len - loc)
        flag=0
        while (text_len != loc):
            if length == 1:
                loc += 1
                length = min(max_len, text_len - loc)
                flag-=1
            elif text[loc:loc + length] in dictlist[length-1]:
                if flag == 2 and merge>0:
                    if jiange==1:
                        uList[-1] = unity(uList[-1].name +'/'+ text[loc:loc + length])
                    else:
                        uList[-1] = unity(uList[-1].name + text[loc:loc + length])
                elif flag==1 and merge==2:
                    if '\u4e00' <= text[loc-1] <= '\u9fff':
                        if jiange==1:
                            uList[-1]=unity(uList[-1].name +'/'+ text[loc-1:loc + length])
                        else:
                            uList[-1] = unity(uList[-1].name + text[loc - 1:loc + length])
                    elif text[loc-1] in ['\"','“','”','‘','’']:
                        uList[-1] = unity(uList[-1].name + text[loc:loc + length])
                    else:
                        uList.append(unity(text[loc:loc + length]))
                else:
                    uList.append(unity(text[loc:loc + length]))
                loc += length
                length = min(max_len, text_len - loc)
                flag=2
            else:
                length -= 1
        return uList
    #max_len = max(len(ele) for ele in dict)
    max_len=10
    dict_yusu=[]
    dictlist=[]
    for i in range(max_len):
        dictlist.append([])
    for ele in dict:
        if len(ele)<=max_len:
           dictlist[len(ele)-1].append(ele)
    uList_list=[]
    for i,text in enumerate(text_list):
        #print(i)
        uList_list.append(match(text,max_len,dictlist))

    return uList_list


def csv_to_xlsx(path,file):
    #将文件夹file里的csv文件合并到.xlsx文件path中
    workbook = openpyxl.Workbook()
    for dir in os.listdir(file):
        sheetname=os.path.splitext(dir)[0]
        with open(file+'\\'+dir, 'r', encoding='utf-8-sig') as f:
            read = csv.reader(f)
            sheet = workbook.create_sheet(sheetname)
            l = 1
            for line in read:
                #print(line)
                r = 1
                for i in line:
                    #print(i)
                    sheet.cell(row=l,column=r,value=i)
                    r = r + 1
                l = l + 1
    sheet=workbook['Sheet']
    workbook.remove(sheet)
    workbook.save(path)
def Pos(text):
    seg = posseg.cut(text)
    poslist = gen2list(seg)
    return poslist
def csv2list(path,encoding='utf-8'):
    list=[]
    infile = open(path, 'r',encoding=encoding)
    reader=csv.reader(infile)
    for line in reader:
        list.append(line)
    return list

def sheet2list(sheet):
    List = []
    for i in range(1, sheet.max_row+1):
        List.append([])
        for j in range(1,sheet.max_column+1):
            if sheet.cell(i, j).value!=None:
                List[-1].append(sheet.cell(i, j).value)
            else:
                List[-1].append('')
    return List
def Write2sheet(reslist,sheet):
    for i in range(len(reslist)):
        sheet.append(reslist[i])
pronoun=['这','他','它','其','本','此','她','该']
binglie=['、','或','和']