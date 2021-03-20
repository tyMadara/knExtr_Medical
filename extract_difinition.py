#利用百度文库的资源提取实体的定义
import docx
import re
import openpyxl
from win32com.client import Dispatch
import sys
import os

def extract_difinition(path):
    '''将此地址下所有.docx中的信息提取到xlsx文件'''
    '''输入：文件地址'''
    wb = openpyxl.Workbook()
    for dir in os.listdir(path):
        if os.path.splitext(dir)[1]=='.docx':
            text1, text2 = extract_information(path+'/'+dir)
            text1,text2 = regular(text1,text2)
            for i in range(len(text1)):
                print(text1[i], "   ", text2[i])
            ws = wb.create_sheet(title=os.path.basename(os.path.splitext(dir)[0]))
            i=0
            for entity in text1:
                ws.cell(row=i + 1, column=1, value=i)
                ws.cell(row=i+1,column=2,value=entity)
                ws.cell(row=i+1,column=3,value='解释')
                ws.cell(row=i+1,column=4,value=text2[i])
                i+=1
    wb.save(path+'/'+'名词解释.xlsx')
    wb.close()

def docToDocx(docPath, docxPath):
    '''将doc转存为docx'''
    word = Dispatch('Word.Application')
    pathPrefix = sys.path[0]+'\\'
    doc = word.Documents.Open(pathPrefix+docPath)
    doc.SaveAs(pathPrefix+docxPath, FileFormat=12)
    doc.Close()
    word.Quit()

def extract_information(path):
    '''输入：.docx文件的地址'''
    '''输出：实体列表text1，定义或解释text2'''
    '''按段落和\n分隔字符，然后尝试用不同模式提取信息'''
    text1=[]
    text2=[]
    doc = docx.Document(path)
    print('文件名：',os.path.basename(path))
    print('未录入的段落:')
    i=0
    for p in doc.paragraphs:
        a=p.text.split('\n')
        for text in a:
            text=re.sub(r'\t',"",text).strip()        #为了不去掉英文之间的空格所以不用\s
            matchObj=re.search(r'(.*?)[:：](.*)',text)
            if matchObj==None:
                matchObj=re.search(r'(.*?)[—-―-][—-―-]+(.*)',text)
            if matchObj==None:
                matchObj=re.search(r'(.*?)是指(.*)',text)
            if matchObj == None:
                matchObj = re.search(r'(.*?)代表(.*)', text)
            if matchObj==None:
                matchObj=re.search(r'(.*?[(（].*?[)）])(.*)',text)
            if matchObj==None:
                matchObj=re.search(r'(.*?[\[【].*[\]】])(.*)',text)
            if matchObj!=None and matchObj.group(1)[0]!='第':    #去除‘第一章：xxx’的情况，尽管这可能漏掉一部分
                text1.append(matchObj.group(1))
                text2.append(matchObj.group(2))
                i+=1
            else:
                if text!='':
                    print(text)
                    i+=1
    print('段落总计:',i)
    print('录入条目:',len(text1))
    return text1,text2

def regular(text1,text2):
    '''去掉标号和首位空格'''
    text1=[re.sub(r'^[0-9①②③④⑤]*[.．。、)）]?',"",entity) for entity in text1]
    text1=[entity.strip() for entity in text1]
    text2=[difinition.strip() for difinition in text2]
    return text1,text2


#docToDocx('../名词解释/诊断学3.doc','../名词解释/诊断学3.docx')
#path='../名词解释'
#extract_difinition(path)
